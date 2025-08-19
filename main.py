from random import choice, shuffle, randint
from characteristic import *
from openpyxl import Workbook
import datetime

class Player:
    def __init__(self,
                 number,
                 gender,
                 age,
                 fertility,
                 body_type,
                 trait,
                 profession,
                 experience,
                 health,
                 hobby,
                 phobia,
                 backpack,
                 extra_info,
                 special_ability):
        self.number = number
        self.gender = gender
        self.age = age
        if (self.gender == "Жінка" and self.age > 52) or (self.gender == "Чоловік" and self.age > 60):
            self.fertility = "Неплідний"
        else:
            self.fertility = fertility
        self.body_type = body_type
        self.trait = trait
        self.profession = profession
        self.experience = experience
        self.health = health
        self.hobby = hobby
        self.phobia = phobia
        self.backpack = backpack
        self.extra_info = extra_info
        self.special_ability = special_ability

    def to_list(self):
        return [
            self.number,
            self.gender,
            self.age,
            self.fertility,
            self.body_type,
            self.trait,
            self.experience,
            self.profession,
            self.health,
            self.hobby,
            self.phobia,
            self.backpack,
            self.extra_info,
            self.special_ability
        ]

def generate_health_pool(count):
    healthy_count = choice(range(2))
    pool = ["Абсолютно здоровий"] * healthy_count
    all_diseases = health_states.copy()
    if "Здоровий" in all_diseases:
        all_diseases.remove("Здоровий")
    shuffle(all_diseases)
    remaining_count = count - len(pool)
    pool += all_diseases[:min(remaining_count, len(all_diseases))]
    while len(pool) < count:
        for disease in all_diseases:
            if len(pool) < count:
                pool.append(disease)
            else:
                break
    shuffle(pool)
    return pool

def generate_pool(options, total_count):
    pool = options.copy()
    shuffle(pool)
    while len(pool) < total_count:
        pool += options.copy()
        shuffle(pool)
    return pool[:total_count]

def generate_random_players(count):
    health_pool = generate_health_pool(count)
    trait_pool = generate_pool(human_traits, count)
    profession_pool = generate_pool(professions, count)
    hobby_pool = generate_pool(hobbies, count)
    phobia_pool = generate_pool(fears, count)
    backpack_pool = generate_pool(backpacks, count)
    extra_info_pool = generate_pool(additional_info, count)
    special_ability_pool = generate_pool(special_abilities, count)

    print(
        f"Пул здоров'я: {health_pool}\n"
        f"Пул рис характеру: {trait_pool}\n"
        f"Пул професій: {profession_pool}\n"
        f"Пул хобі: {hobby_pool}\n"
        f"Пул фобій: {phobia_pool}\n"
        f"Пул рюкзаків: {backpack_pool}\n"
        f"Пул додаткової інформації: {extra_info_pool}\n"
        f"Пул спец. можливостей: {special_ability_pool}"
    )

    players = []
    for i in range(count):
        gender = choice(genders)
        age = randint(18, 80)
        fertility = choice(["Плідний", "Неплідний"])
        experience = randint(0, max(0, age - 18))
        player = Player(
            number=i+1,
            gender=gender,
            age=age,
            fertility=fertility,
            body_type=choice(body_types),
            trait=trait_pool[i],
            profession=profession_pool[i],
            experience=experience,
            health=health_pool[i],
            hobby=hobby_pool[i],
            phobia=phobia_pool[i],
            backpack=backpack_pool[i],
            extra_info=extra_info_pool[i],
            special_ability=special_ability_pool[i]
        )
        players.append(player)
    return players


def generate_players_excel_vertical(players):
    wb = Workbook()
    ws = wb.active
    ws.title = "Гравці"
    characteristics = [
        "№", "Стать", "Вік", "Плідність", "Статура", "Риса характеру",
        "Стаж", "Професія", "Здоров’я", "Хобі", "Фобія",
        "Рюкзак", "Додаткове", "Спец. можливість"
    ]
    for i, char in enumerate(characteristics, start=1):
        ws.cell(row=i, column=1, value=char)

    for j, player in enumerate(players):
        data = player.to_list()
        for i, value in enumerate(data, start=1):
            ws.cell(row=i, column=j + 2, value=value)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    filename = f"players_vertical_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"Файл '{filename}' створено!")

if __name__ == "__main__":
    n = int(input("Введіть кількість гравців для генерації: "))
    players = generate_random_players(n)
    generate_players_excel_vertical(players)
